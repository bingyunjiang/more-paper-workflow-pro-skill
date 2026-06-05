#!/usr/bin/env python3
"""Generate full deliverable set from 检索文献表.md — the ONE command for Step 4g.

Produces from a single .md input:
  1. 检索文献表.xlsx  — openpyxl workbook with filter/sort, frozen header
  2. 文献库.bib        — BibTeX export with Tier/Score/influential_citations in note
  3. 检索文献表.pdf    — via md_to_pdf.py (only with --pdf flag)

Usage:
  # Standard (xlsx + bib only)
  python3 scripts/generate_retrieval_report.py 检索文献表.md

  # Also generate PDF
  python3 scripts/generate_retrieval_report.py 检索文献表.md --pdf

  # Custom output basename
  python3 scripts/generate_retrieval_report.py 检索文献表.md --basename my_literature

Output files are written alongside the input .md file unless --outdir is specified.
"""

import argparse
import json
import os
import re
import subprocess
import sys
import textwrap
from pathlib import Path


# ── Markdown table parser ──────────────────────────────────────────────

def _normalize_doi(doi: str) -> str:
    """Strip whitespace, unify prefix."""
    doi = doi.strip()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if doi.lower().startswith(prefix):
            doi = doi[len(prefix):]
    return doi


def _parse_md_table(md_path: str) -> list[dict]:
    """Parse a markdown table from 检索文献表.md into a list of dicts.

    Returns list of rows, each a dict with keys matching the table columns.
    Handles tables with columns: DOI, Title/标题, Year/年份, Source/来源,
    Score/评分, Tier, Flags/旗标, Citations/引用, influential_citations, Sub-topic/子课题.
    """
    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    rows = []

    # Find markdown tables: lines starting with | that contain column separators
    # Look for a header row followed by a separator row
    lines = text.split("\n")

    # Find all table blocks
    table_blocks = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("|") and "|" in line[1:]:
            # Check if next line is a separator (contains :--- or ---)
            if i + 1 < len(lines) and re.match(r'^\|[\s\-:|\s]+\|', lines[i + 1].strip()):
                # Found a table: header at i, separator at i+1
                header_line = line
                # Collect all data rows
                data_rows = []
                j = i + 2
                while j < len(lines) and lines[j].strip().startswith("|"):
                    data_rows.append(lines[j].strip())
                    j += 1
                table_blocks.append((header_line, data_rows))
                i = j
                continue
        i += 1

    if not table_blocks:
        print("⚠️  No markdown tables found in input file", file=sys.stderr)
        return rows

    # Use the LARGEST table (most columns) — this is the main literature table
    best_header = ""
    best_data = []
    best_cols = 0
    for header_line, data_rows in table_blocks:
        cols = [c.strip() for c in header_line.strip("|").split("|")]
        if len(cols) > best_cols:
            best_cols = len(cols)
            best_header = header_line
            best_data = data_rows

    if not best_header:
        return rows

    columns = [c.strip().lower() for c in best_header.strip("|").split("|")]

    # Build column index mapping (flexible matching)
    col_map = {}
    for idx, col in enumerate(columns):
        col_clean = col.strip().lower()
        if "doi" in col_clean:
            col_map["doi"] = idx
        elif "title" in col_clean or "标题" in col_clean or "题目" in col_clean:
            col_map["title"] = idx
        elif "year" in col_clean or "年份" in col_clean or "年" in col_clean:
            col_map["year"] = idx
        elif "source" in col_clean or "来源" in col_clean:
            col_map["source"] = idx
        elif "score" in col_clean or "评分" in col_clean or "分数" in col_clean:
            col_map["score"] = idx
        elif "tier" in col_clean:
            col_map["tier"] = idx
        elif "flag" in col_clean or "旗标" in col_clean or "标记" in col_clean:
            col_map["flags"] = idx
        elif "citation" in col_clean or "引用" in col_clean:
            if "influential" in col_clean or "影响力" in col_clean or "高引" in col_clean:
                col_map["influential_citations"] = idx
            else:
                col_map["citations"] = idx
        elif "sub" in col_clean or "子课题" in col_clean or "topic" in col_clean:
            col_map["subtopic"] = idx
        elif "author" in col_clean or "作者" in col_clean:
            col_map["authors"] = idx
        elif "journal" in col_clean or "期刊" in col_clean or "venue" in col_clean:
            col_map["journal"] = idx

    # Parse data rows
    for row_line in best_data:
        cells = [c.strip() for c in row_line.strip("|").split("|")]
        row = {}

        for key, idx in col_map.items():
            if idx < len(cells):
                row[key] = cells[idx]
            else:
                row[key] = ""

        # Clean DOI
        if "doi" in row:
            row["doi"] = _normalize_doi(row["doi"])

        if row.get("doi"):
            rows.append(row)

    print(f"📋 Parsed {len(rows)} papers from markdown table ({len(columns)} columns)", flush=True)
    return rows


def _escape_bibtex(s: str) -> str:
    """Escape special characters for BibTeX."""
    if not s:
        return ""
    # Replace LaTeX special characters
    for char, repl in [("&", "\\&"), ("%", "\\%"), ("$", "\\$"),
                       ("#", "\\#"), ("_", "\\_"), ("{", "\\{"), ("}", "\\}"),
                       ("~", "\\~{}"), ("^", "\\^{}")]:
        s = s.replace(char, repl)
    return s


def _bibtex_key(authors_str: str, year: str, title: str) -> str:
    """Generate a BibTeX citation key: FirstAuthorYear_FirstWord."""
    first_author = authors_str.split(",")[0].split(" and ")[0].strip() if authors_str else "Unknown"
    # Take last name (last word after stripping)
    last_name = first_author.split()[-1] if first_author.split() else first_author
    # Clean: remove non-alphanumeric
    last_name = re.sub(r'[^a-zA-Z0-9]', '', last_name)
    year_str = str(year).strip() if year else "????"
    first_word = ""
    if title:
        words = title.strip().split()
        for w in words:
            clean = re.sub(r'[^a-zA-Z0-9]', '', w)
            if clean and len(clean) > 2:
                first_word = clean.lower()
                break
    if not first_word:
        first_word = "unknown"
    return f"{last_name}{year_str}_{first_word}"


# ── Excel generation ───────────────────────────────────────────────────

def _generate_xlsx(rows: list[dict], output_path: str):
    """Generate .xlsx from parsed rows using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "检索文献表"

    # Define columns (order matters)
    columns = [
        ("doi", "DOI"),
        ("title", "标题"),
        ("authors", "作者"),
        ("year", "年份"),
        ("journal", "期刊/会议"),
        ("source", "来源"),
        ("score", "评分"),
        ("tier", "Tier"),
        ("citations", "引用数"),
        ("influential_citations", "影响力引用"),
        ("flags", "旗标"),
        ("subtopic", "子课题"),
    ]

    # Header styling
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data styling
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=False)

    # Tier color fills
    tier_fills = {
        "T1": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
        "T2": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # blue
        "T3": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # orange
    }

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Write header
    for col_idx, (_, header_label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Color-code Tier column
            if key == "tier" and value in tier_fills:
                cell.fill = tier_fills[value]

    # Column widths
    col_widths = {
        "doi": 38, "title": 60, "authors": 25, "year": 7,
        "journal": 30, "source": 14, "score": 7, "tier": 7,
        "citations": 10, "influential_citations": 13, "flags": 12, "subtopic": 14,
    }
    for col_idx, (key, _) in enumerate(columns, start=1):
        width = col_widths.get(key, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    wb.save(output_path)
    print(f"✅ Generated {output_path} ({len(rows)} rows)", flush=True)
    return True


# ── BibTeX generation ──────────────────────────────────────────────────

def _generate_bibtex(rows: list[dict], output_path: str):
    """Generate .bib from parsed rows."""
    lines = []
    lines.append(f"% Generated by generate_retrieval_report.py")
    lines.append(f"% {len(rows)} references (T1-T3 only)")
    lines.append("")

    for row in rows:
        authors_raw = row.get("authors", "")
        title = row.get("title", "Unknown Title")
        year = row.get("year", "????")
        doi = row.get("doi", "")
        journal = row.get("journal", "")
        source = row.get("source", "")
        tier = row.get("tier", "")
        score = row.get("score", "")
        influential = row.get("influential_citations", "")
        subtopic = row.get("subtopic", "")

        cite_key = _bibtex_key(authors_raw, year, title)

        # Build note field
        note_parts = []
        if tier:
            note_parts.append(f"Tier {tier}")
        if score:
            note_parts.append(f"Score: {score}")
        if source:
            note_parts.append(f"source: {source}")
        if influential:
            note_parts.append(f"influential_citations: {influential}")
        if subtopic:
            note_parts.append(f"subtopic: {subtopic}")
        note = " | ".join(note_parts)

        # Author formatting: "Last, First and Last, First"
        author_list = [a.strip() for a in authors_raw.split(";") if a.strip()] if authors_raw else []
        if not author_list:
            author_list = [authors_raw] if authors_raw else ["Unknown"]
        author_str = " and ".join(author_list)

        lines.append(f"@article{{{cite_key},")
        lines.append(f"  title     = {{{_escape_bibtex(title)}}},")
        lines.append(f"  author    = {{{_escape_bibtex(author_str)}}},")
        if journal:
            lines.append(f"  journal   = {{{_escape_bibtex(journal)}}},")
        lines.append(f"  year      = {{{year}}},")
        if doi:
            lines.append(f"  doi       = {{{doi}}},")
        if note:
            lines.append(f"  note      = {{{_escape_bibtex(note)}}},")
        lines.append("}")
        lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"✅ Generated {output_path} ({len(rows)} references)", flush=True)
    return True


# ── PDF generation ─────────────────────────────────────────────────────

def _generate_pdf(md_path: str, output_path: str) -> bool:
    """Generate .pdf from .md using md_to_pdf.py."""
    script_dir = Path(__file__).resolve().parent
    md_to_pdf = script_dir / "md_to_pdf.py"

    if not md_to_pdf.exists():
        print(f"⚠️  md_to_pdf.py not found at {md_to_pdf}, skipping PDF generation",
              file=sys.stderr)
        return False

    cmd = [
        sys.executable, str(md_to_pdf),
        md_path,
        "-o", output_path,
        "--type", "literature_table",
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode == 0:
            print(f"✅ Generated {output_path}", flush=True)
            return True
        else:
            print(f"⚠️  PDF generation failed: {result.stderr.strip()}", file=sys.stderr)
            # Try without --type flag
            cmd2 = [sys.executable, str(md_to_pdf), md_path, "-o", output_path]
            result2 = subprocess.run(cmd2, capture_output=True, text=True, timeout=120)
            if result2.returncode == 0:
                print(f"✅ Generated {output_path} (fallback mode)", flush=True)
                return True
            print(f"⚠️  PDF generation also failed in fallback mode", file=sys.stderr)
            return False
    except subprocess.TimeoutExpired:
        print(f"⚠️  PDF generation timed out", file=sys.stderr)
        return False
    except Exception as e:
        print(f"⚠️  PDF generation error: {e}", file=sys.stderr)
        return False


# ── Saturation summary extraction ──────────────────────────────────────

def _extract_saturation_summary(md_path: str) -> dict | None:
    """Extract saturation summary from the .md header block."""
    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    info = {}
    # Extract key metrics from the 检索概况 block
    patterns = {
        "search_date": r"检索日期[：:]\s*(.+)",
        "tier": r"Tier[：:]\s*(.+)",
        "databases": r"数据库[：:]\s*(.+)",
        "final_count": r"最终文献[：:]\s*(.+?)(?:\n|$)",
        "coverage": r"饱和度[：:]\s*(.+?)(?:\n|$)",
    }
    for key, pat in patterns.items():
        m = re.search(pat, text)
        if m:
            info[key] = m.group(1).strip()

    return info if info else None


# ── Main ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate full deliverable set (xlsx + bib + pdf) from 检索文献表.md",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
        Examples:
          %(prog)s 检索文献表.md
              → 检索文献表.xlsx + 文献库.bib

          %(prog)s 检索文献表.md --pdf
              → 检索文献表.xlsx + 文献库.bib + 检索文献表.pdf

          %(prog)s 检索文献表.md --basename my_paper_literature
              → my_paper_literature.xlsx + my_paper_literature.bib
        """),
    )
    parser.add_argument("md_path", help="Path to 检索文献表.md")
    parser.add_argument("--pdf", action="store_true", help="Also generate PDF via md_to_pdf.py (opt-in).")
    parser.add_argument("--basename", help="Custom basename for output files (default: derived from input)")
    parser.add_argument("--outdir", help="Output directory (default: same as input .md)")
    args = parser.parse_args()

    md_path = Path(args.md_path).resolve()
    if not md_path.exists():
        print(f"❌ Input file not found: {md_path}", file=sys.stderr)
        sys.exit(1)

    # Determine output directory and basename
    outdir = Path(args.outdir).resolve() if args.outdir else md_path.parent
    basename = args.basename or md_path.stem  # e.g. "检索文献表"

    outdir.mkdir(parents=True, exist_ok=True)
    xlsx_path = outdir / f"{basename}.xlsx"
    bib_path = outdir / f"{'文献库' if basename == '检索文献表' else basename + '_文献库'}.bib"
    pdf_path = outdir / f"{basename}.pdf"

    print(f"\n{'='*60}")
    print(f"📦 Step 4g: 生成检索报告全套交付物")
    print(f"{'='*60}")
    print(f"   输入: {md_path}")
    print(f"   输出目录: {outdir}")
    print()

    # 1. Parse the markdown table
    print("─" * 40)
    print("📋 1/4 解析检索文献表...")
    rows = _parse_md_table(str(md_path))

    if not rows:
        print("❌ No data rows found in markdown table. Cannot generate deliverables.",
              file=sys.stderr)
        sys.exit(1)

    # Show summary
    tier_counts = {}
    for r in rows:
        t = r.get("tier", "?").strip()
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"   T1: {tier_counts.get('T1', 0)}  |  T2: {tier_counts.get('T2', 0)}  |  T3: {tier_counts.get('T3', 0)}")
    print()

    # 2. Generate .xlsx
    print("─" * 40)
    print("📊 2/4 生成 Excel 检索文献表...")
    xlsx_ok = _generate_xlsx(rows, str(xlsx_path))
    print()

    # 3. Generate .bib
    print("─" * 40)
    print("📚 3/4 生成 BibTeX 文献库...")
    bib_ok = _generate_bibtex(rows, str(bib_path))
    print()

    # 4. Generate .pdf (opt-in only)
    pdf_ok = False
    if args.pdf:
        print("─" * 40)
        print("📄 4/4 生成 PDF 检索文献表...")
        pdf_ok = _generate_pdf(str(md_path), str(pdf_path))
        print()
    else:
        print("⏭️  4/4 PDF generation skipped (use --pdf to enable)")
        print()

    # 5. Summary
    print("=" * 60)
    print("📦 交付物清单:")
    print(f"   ✅ 检索文献表.md   — {md_path}")
    print(f"   {'✅' if xlsx_ok else '❌'} 检索文献表.xlsx — {xlsx_path}")
    print(f"   {'✅' if bib_ok else '❌'} 文献库.bib     — {bib_path}")
    if args.pdf:
        print(f"   {'✅' if pdf_ok else '❌'} 检索文献表.pdf  — {pdf_path}")

    # Extract and show saturation summary
    sat = _extract_saturation_summary(str(md_path))
    if sat:
        print()
        print("📈 检索概况 (从 .md 提取):")
        for k, v in sat.items():
            print(f"   {k}: {v}")

    print()
    all_ok = xlsx_ok and bib_ok
    if args.pdf:
        all_ok = all_ok and pdf_ok

    if all_ok:
        print("🎉 全套交付物生成完成！")
    else:
        print("⚠️  部分交付物生成失败，请检查上方输出。")
        sys.exit(1)


if __name__ == "__main__":
    main()
